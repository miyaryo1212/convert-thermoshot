import ctypes
import datetime
import os
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import ttk

import matplotlib
import numpy as np
import openpyxl as opxl


def showmsgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)


def selectmode():
    root = tkinter.Tk()
    root.withdraw()

    root.title("select-mode")
    root.geometry("300x130")

    def ok_get(event):
        root.quit()

    var = tkinter.IntVar()
    var.set(0)

    radiobutton_rtp = tkinter.Radiobutton(
        root, value=0, variable=var, text="shift-time"
    )
    radiobutton_tsp = tkinter.Radiobutton(
        root, value=1, variable=var, text="real-time"
    )
    radiobutton_rtp.grid(column=0, row=0, padx=20, pady=10)
    radiobutton_tsp.grid(column=0, row=1, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.bind("<Button-1>", ok_get)
    button.grid(column=1, row=2, padx=20, pady=10)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if var.get() == None:
        print("ERROR: Plotting mode was not select correctly")
        showmsgbox("[ERROR] select-mode", "プロットモードが正しく選択されませんでした")
        quit()

    return var.get()


def selectfile():
    root = tkinter.Tk()
    root.withdraw()
    fType = [("", ".xlsx")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    tkinter.messagebox.showinfo("compare-sheets", "Excelファイル (.xlsx形式) を選択")
    file = tkinter.filedialog.askopenfilename(filetypes=fType, initialdir=iDir)

    if not file:
        tkinter.messagebox.showinfo(
            "[Error] compare-sheets", "Excelファイル (.xlsx形式) が選択されませんでした"
        )
        quit()

    return file


def detectsheets(booksrc):
    workbook = opxl.load_workbook(booksrc)
    sheets = workbook.sheetnames

    return sheets


def selectsheets(list):
    root = tkinter.Tk()
    root.withdraw()

    root.title("select-sheets")
    root.geometry("300x130")

    def ok_get(event):
        root.quit()

    label_before = tkinter.Label(root, text="Before")
    label_before.grid(column=0, row=0, padx=20, pady=10)
    label_after = tkinter.Label(root, text="After")
    label_after.grid(column=0, row=1, padx=20, pady=10)

    combo_before = ttk.Combobox(root, values=list, justify="center")
    combo_before.grid(column=1, row=0, padx=20, pady=10)
    combo_after = ttk.Combobox(root, values=list, justify="center")
    combo_after.grid(column=1, row=1, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.bind("<Button-1>", ok_get)
    button.grid(column=1, row=2, padx=20, pady=10)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not combo_before.get() or not combo_after.get():
        print("ERROR: Sheets were not select correctly")
        showmsgbox("[Error] select-sheets", "シートが正しく選択されませんでした")
        quit()

    return combo_before.get(), combo_after.get()


def readsheet(booksrc, sheet):
    workbook = opxl.load_workbook(booksrc)
    sheet = workbook[sheet]

    list = np.zeros((sheet.max_row, sheet.max_column))
    for row_num in sheet.rows:
        row = row_num[0].row
        for column_num in sheet.columns:
            column = column_num[0].column

            list[row - 1][column - 1] = sheet.cell(
                row=row, column=column
            ).value

    return list


def writesheet(bookpath, sheetname, list):
    workbook = opxl.Workbook()
    sheet = workbook.create_sheet(sheetname)

    for y in range(len(list)):
        for x in range(len(list[0])):
            list_value = list[y][x]
            sheet.cell(y + 1, x + 1, list_value)

    workbook.save(bookpath)

    return None


def show2dplot(mode, list1, list2, name):
    if mode == 0:
        matplotlib.use("Agg")
    else:
        pass

    import matplotlib.pyplot as plt

    print("INFO: generating plot...")

    fig = plt.figure(figsize=(40, 30))
    ax = fig.add_subplot(111)

    ax.set_xlabel("x")
    ax.set_ylabel("y")
    ax.set_title(name)

    width = 1

    for i in range(len(list1)):
        ax.bar(i, list1[i], width=width, color="#FF4B00")
        ax.bar(i, list2[i], width=width, color="#005AFF", bottom=list1[i])

    print("INFO: writing plot...")
    if mode == 1:
        plt.show()
    else:
        pass

    plt.savefig("./saved/figure_{}-{}.png".format(sheets[1], sheets[0]))
    print(
        "INFO: Saved figure image to ./saved/figure2d_{}-{}.png".format(
            sheets[1], sheets[0]
        )
    )

    return None


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    mode = selectmode()

    src = selectfile()
    print("INFO: {} loading file...".format(src))

    sheets = selectsheets(detectsheets(src))
    print(
        "INFO: {} (After), {} (Before) (2 sheets selected)".format(
            sheets[1], sheets[0]
        )
    )

    print("INFO: {} converting sheet...".format(sheets[0]))
    list_before = readsheet(src, sheets[0])
    print("INFO: {} converted sheet".format(sheets[0]))

    print("INFO: {} converting sheet...".format(sheets[1]))
    list_after = readsheet(src, sheets[1])
    print("INFO: {} converted sheet".format(sheets[1]))

    list_delta = list_after - list_before
    print("INFO: transposing 2darray...")
    list_delta = list_delta.T.tolist()
    list_shpere = []
    list_rail = []

    for x in range(len(list_delta)):
            list_shpere.append(sum(list_delta[x][:149]))
            list_rail.append(sum(list_delta[x][:150]))

    print("INFO: transposed 2darray")

    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    dir = "./saved"
    if not os.path.exists(dir):
        os.makedirs(dir)
    bookpath = "./saved/runned_{}.xlsx".format(time_now)
    writesheet(bookpath, ("{}-{}".format(sheets[1], sheets[0])), list_delta)
    print("INFO: Saved workbook to {}".format(bookpath))
    show2dplot(
        mode, list_rail, list_shpere, "{}-{}".format(sheets[1], sheets[0])
    )
    figurepath = "./saved/figure_{}-{}.png".format(sheets[1], sheets[0])

    print("INFO: All operations have been completed")
    showmsgbox(
        "compare-sheets",
        "ワークシート：\t{}\nグラフ：\t\t{}\nに保存済み".format(bookpath, figurepath),
    )
