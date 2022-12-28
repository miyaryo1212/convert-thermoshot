import ctypes
import datetime
import os
import pathlib
import time
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import ttk

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl as opxl
from mpl_toolkits.mplot3d import Axes3D
from openpyxl.utils import get_column_letter


def showmsgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)


def selectfile():
    root = tkinter.Tk()
    root.withdraw()
    fType = [("", ".xlsx")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    tkinter.messagebox.showinfo("compare-sheets", "Excelファイル (.xlsx形式) を選択")
    file = tkinter.filedialog.askopenfilename(filetypes=fType, initialdir=iDir)

    if not file:
        tkinter.messagebox.showinfo(
            "[Error] compare-sheets", "Excelファイル (.xlsx形式) が選択されませんでした"
        )
        quit()

    return file


def detectsheets(booksrc):
    workbook = opxl.load_workbook(booksrc)
    sheets = workbook.sheetnames

    return sheets


def selectsheets(list):
    root = tkinter.Tk()
    root.withdraw()

    root.title("select-sheets")
    root.geometry("300x130")

    def ok_get(event):
        root.quit()

    label_before = tkinter.Label(root, text="Before")
    label_before.grid(column=0, row=0, padx=20, pady=10)
    label_after = tkinter.Label(root, text="After")
    label_after.grid(column=0, row=1, padx=20, pady=10)

    combo_before = ttk.Combobox(root, values=list, justify="center")
    combo_before.grid(column=1, row=0, padx=20, pady=10)
    combo_after = ttk.Combobox(root, values=list, justify="center")
    combo_after.grid(column=1, row=1, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.bind("<Button-1>", ok_get)
    button.grid(column=1, row=2, padx=20, pady=10)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not combo_before.get() or not combo_after.get():
        print("ERROR: Sheets were not selected correctly")
        showmsgbox("[Error] select-sheets", "シートが正しく選択されませんでした")
        quit()

    return combo_before.get(), combo_after.get()


def readsheet(booksrc, sheet):
    workbook = opxl.load_workbook(booksrc)
    sheet = workbook[sheet]

    list = np.zeros((sheet.max_row, sheet.max_column))
    for row_num in sheet.rows:
        row = row_num[0].row
        for column_num in sheet.columns:
            column = column_num[0].column

            list[row - 1][column - 1] = sheet.cell(
                row=row, column=column
            ).value

    return list


def writesheet(bookpath, sheetname, list):
    workbook = opxl.Workbook()
    sheet = workbook.create_sheet(sheetname)

    for y in range(len(list)):
        for x in range(len(list[0])):
            list_value = list[y][x]
            sheet.cell(y + 1, x + 1, list_value)

    workbook.save(bookpath)

    return None


def showplot(list):
    x, y, z = 0, 0, 0
    dx, dy, dz = 1, 1, 1

    fig = plt.figure(figsize=(9, 8))
    ax = fig.add_subplot(111, projection="3d")

    top = x + y
    bottom = np.zeros_like(top)
    width = depth = 1

    ax.set_xlabel("x")
    ax.set_ylabel("y")
    ax.set_zlabel("z")
    ax.set_title("No Title")

    for y in range(len(list)):
        for x in range(len(list[0])):
            z = list[y][x]
            ax.bar3d(x, y, bottom, width, depth, z, shade=True)

    plt.show()

    return None


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    src = selectfile()
    print("INFO: {} loading file...".format(src))

    sheets = selectsheets(detectsheets(src))
    print("INFO: {}, {} (2 sheets selected)".format(sheets[0], sheets[1]))

    print("INFO: {} converting sheet...".format(sheets[0]))
    list_before = readsheet(src, sheets[0])
    print("INFO: {} converted sheet".format(sheets[0]))

    print("INFO: {} converting sheet...".format(sheets[1]))
    list_after = readsheet(src, sheets[1])
    print("INFO: {} converted sheet".format(sheets[1]))

    list_delta = list_after - list_before

    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    dir = "./saved"
    if not os.path.exists(dir):
        os.makedirs(dir)
    bookpath = "./saved/runned_{}.xlsx".format(time_now)
    writesheet(bookpath, ("{}-{}".format(sheets[1], sheets[0])), list_delta)
    print("INFO: preparing plot...")
    showplot(list_delta)
