import ctypes
import datetime
import os
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import ttk

import numpy as np
import openpyxl as opxl


def showmsgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)


def selectmode():
    root = tkinter.Tk()
    root.withdraw()

    root.title("select-mode")
    root.geometry("300x130")

    def ok_get(event):
        root.quit()

    var = tkinter.IntVar()
    var.set(0)

    radiobutton_rtp = tkinter.Radiobutton(
        root, value=0, variable=var, text="read existing file"
    )
    radiobutton_tsp = tkinter.Radiobutton(
        root, value=1, variable=var, text="generate new file"
    )
    radiobutton_rtp.grid(column=0, row=0, padx=20, pady=10)
    radiobutton_tsp.grid(column=0, row=1, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.bind("<Button-1>", ok_get)
    button.grid(column=1, row=2, padx=20, pady=10)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if var.get() == None:
        print("ERROR: Plotting mode was not select correctly")
        showmsgbox("[ERROR] select-mode", "読み込みモードが正しく選択されませんでした")
        quit()

    print("INFO: mode {} selected".format(var.get()))
    return var.get()


def askfromto():
    root = tkinter.Tk()
    root.withdraw()

    root.title("ask-interval")
    root.geometry("300x200")

    def ok_get(event):
        root.quit()

    start_label = tkinter.Label(root, text="Start (s)")
    start_label.grid(column=0, row=0, padx=20, pady=10)
    end_label = tkinter.Label(root, text="End (s)")
    end_label.grid(column=0, row=1, padx=20, pady=10)

    start_time = tkinter.Entry(root, width=20)
    start_time.grid(column=1, row=0, padx=20, pady=10)
    end_time = tkinter.Entry(root, width=20)
    end_time.grid(column=1, row=1, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.grid(column=1, row=3, padx=20, pady=10)
    button.bind("<Button-1>", ok_get)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not start_time.get() or not end_time.get():
        print("ERROR: Start/End time were not inputed correctly")
        showmsgbox("[ERROR] ask-fromto", "開始／終了時刻が正しく入力されませんでした")
        quit()

    print("INFO: input {}s >> {}s".format(start_time.get(), end_time.get()))
    return float(start_time.get()), float(end_time.get())


def selecttime(time):
    root = tkinter.Tk()
    root.withdraw()

    root.title("select-time")
    root.geometry("300x150")

    def ok_get(event):
        root.quit()

    label = tkinter.Label(root, text="time")
    label.grid(column=0, row=0, padx=20, pady=10)

    spinbox = tkinter.Spinbox(root, from_=time[0], to=time[1], increment=0.01)
    spinbox.grid(column=1, row=0, padx=20, pady=10)

    button = tkinter.Button(root, text="OK")
    button.grid(column=1, row=2, padx=20, pady=10)
    button.bind("<Button-1>", ok_get)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not spinbox.get():
        print("ERROR time was not input correctly")
        showmsgbox("[ERROR] select-time", "時間が正しく入力されませんでした ")
        quit()

    print("INFO: {}s time selected".format(spinbox.get()))
    return float(spinbox.get())


def selectfile(mode):
    root = tkinter.Tk()
    root.withdraw()
    fType = [("", ".xlsx")]
    iDir = os.path.abspath(os.path.dirname(__file__))

    if mode == 0:
        tkinter.messagebox.showinfo(
            "compare-sheets", "生成済み.xlsx形式Excelファイルを選択"
        )
    else:
        tkinter.messagebox.showinfo(
            "compare-sheets", "解析する.xlsx形式Excelファイルを選択"
        )

    src = tkinter.filedialog.askopenfilename(filetypes=fType, initialdir=iDir)

    if not src:
        tkinter.messagebox.showinfo(
            "[Error] compare-sheets", ".xlsx形式Excelファイルが正しく選択されませんでした"
        )
        quit()

    print("INFO: {} book selected".format(src))
    return src


def selectsheets(booksrc, mode):
    workbook = opxl.load_workbook(booksrc)
    sheets = workbook.sheetnames

    root = tkinter.Tk()
    root.withdraw()

    if mode == 0:
        root.title("select-sheet")
        root.geometry("300x130")

        def ok_get(event):
            root.quit()

        label = tkinter.Label(root, text="sheet")
        label.grid(column=0, row=0, padx=20, pady=10)

        combo = ttk.Combobox(root, values=sheets, justify="center")
        combo.grid(column=1, row=0, padx=20, pady=10)

        button = tkinter.Button(root, text="OK")
        button.grid(column=1, row=1, padx=20, pady=10)
        button.bind("<Button-1>", ok_get)

        root.bind("<Return>", ok_get)

        root.deiconify()
        root.mainloop()
        root.withdraw()

        if not combo.get():
            print("ERROR: Sheet was not select correctly")
            showmsgbox("[ERROR] select-sheet", "シートが正しく選択されませんでした")
            quit()

        print("INFO: {} (Delta) *1 sheet selected".format(combo.get()))
        return combo.get()

    else:
        root.title("select-sheets")
        root.geometry("300x130")

        def ok_get(event):
            root.quit()

        label_before = tkinter.Label(root, text="Before")
        label_before.grid(column=0, row=0, padx=20, pady=10)
        label_after = tkinter.Label(root, text="After")
        label_after.grid(column=0, row=1, padx=20, pady=10)

        combo_before = ttk.Combobox(root, values=sheets, justify="center")
        combo_before.grid(column=1, row=0, padx=20, pady=10)
        combo_after = ttk.Combobox(root, values=sheets, justify="center")
        combo_after.grid(column=1, row=1, padx=20, pady=10)

        button = tkinter.Button(root, text="OK")
        button.grid(column=1, row=2, padx=20, pady=10)
        button.bind("<Button-1>", ok_get)

        root.bind("<Return>", ok_get)

        root.deiconify()
        root.mainloop()
        root.withdraw()

        if not combo_before.get() or not combo_after.get():
            print("ERROR: Sheets were not select correctly")
            showmsgbox("[Error] select-sheets", "シートが正しく選択されませんでした")
            quit()

        print(
            "INFO: {} (Before) {} (After) *2 sheets selected".format(
                combo_before.get(), combo_after.get()
            )
        )
        return combo_before.get(), combo_after.get()


def readsheet(booksrc, sheet):
    print("INFO: {} reading a sheet in {} book...".format(sheet, booksrc))
    workbook = opxl.load_workbook(booksrc)
    sheet = workbook[sheet]

    list = np.zeros((sheet.max_row, sheet.max_column))
    for row_num in sheet.rows:
        row = row_num[0].row
        for column_num in sheet.columns:
            column = column_num[0].column

            list[row - 1][column - 1] = sheet.cell(
                row=row, column=column
            ).value

    print("INFO: {} read a sheet in {} book".format(sheet, booksrc))
    return list


def writesheet(bookpath, sheetname, list, timefromto, timeselected):
    workbook = opxl.Workbook()
    sheet = workbook.create_sheet(sheetname)

    for y in range(len(list)):
        for x in range(len(list[0])):
            list_value = (
                (list[y][x])
                * (timeselected - timefromto[0])
                / (timefromto[1] - timefromto[0])
            )
            sheet.cell(y + 1, x + 1, list_value)
            pass

    workbook.save(bookpath)

    return None


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    mode = selectmode()
    booksrc = selectfile(mode)
    time_fromto = askfromto()
    time_selected = selecttime(time_fromto)

    if mode == 0:
        sheet = selectsheets(booksrc, mode)
        sheetname = "{}_{}s".format(sheet, time_selected)
        delta = readsheet(booksrc, sheet)
    else:
        sheets = selectsheets(booksrc, mode)
        before = readsheet(booksrc, sheets[0])
        after = readsheet(booksrc, sheets[1])
        sheetname = "{}-{}_{}s".format(sheets[1], sheets[0], time_selected)
        delta = after - before

    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    dir = "./saved"
    if not os.path.exists(dir):
        os.makedirs(dir)
    bookpath = "./saved/runned_{}.xlsx".format(time_now)

    writesheet(bookpath, sheetname, delta, time_fromto, time_selected)

    print("INFO: All operations have been completed")
    showmsgbox("complement-sheets", "{} に保存済み".format(bookpath))
