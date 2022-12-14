import ctypes
import datetime
import os
import pathlib
import tkinter
import tkinter.filedialog
import tkinter.messagebox

import cv2 as cv
import openpyxl as opxl
from openpyxl.utils import get_column_letter


def showmsgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)


def choosefiles():
    root = tkinter.Tk()
    root.withdraw()
    fType = [("", "*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    tkinter.messagebox.showinfo(
        "convert-thermoshot", "F30サーモショットで撮影したグレースケール画像を選択"
    )
    files = tkinter.filedialog.askopenfilenames(
        filetypes=fType, initialdir=iDir
    )

    if not files:
        tkinter.messagebox.showinfo(
            "[Error] convert-thermoshot", "画像が選択されませんでした"
        )
        quit()

    files = list(files)
    # print(files)

    return files


def askminmax():
    root = tkinter.Tk()
    root.withdraw()

    root.title("設定温度を小数第1位まで入力")
    root.geometry("400x120")

    def ok_get(event):
        root.quit()

    label_min = tkinter.Label(root, text="最小温度：")
    label_max = tkinter.Label(root, text="最大温度：")
    label_min.place(x=20, y=20)
    label_max.place(x=20, y=50)

    editbox_min = tkinter.Entry(root, width=40)
    editbox_min.insert(tkinter.END, "")
    editbox_min.place(x=100, y=20)

    editbox_max = tkinter.Entry(root, width=40)
    editbox_max.insert(tkinter.END, "")
    editbox_max.place(x=100, y=50)

    button = tkinter.Button(root, text="OK")
    button.bind("<Button-1>", ok_get)
    button.place(x=300, y=80)

    root.bind("<Return>", ok_get)

    root.deiconify()
    root.mainloop()
    root.withdraw()

    if not (editbox_min.get() or editbox_max.get()):
        tkinter.messagebox.showinfo("convert-thermoshot", "設定温度が正しく入力されませんでした")
        quit()

    min, max = float(editbox_min.get()), float(editbox_max.get())

    return min, max


def rdimg(src, min, max):
    print("INFO: {} loading".format(src))
    img = cv.imread(src)
    img_hsv = cv.cvtColor(img, cv.COLOR_BGR2HSV)
    h, s, v = cv.split(img_hsv)

    print("INFO: {} processing...".format(src))
    for y in range(len(v)):
        for x in range(len(v[0])):
            v[y][x] = min + ((max - min) / 255 * v[y][x])

    # cv.imshow("convert-thermoshot - Press Esc to continue", img)
    # cv.imshow("img hsv", img_hsv)
    # cv.waitKey(0)
    # cv.destroyAllWindows()
    print("INFO: {} finished".format(src))

    return v


def cvtsheet(workbook, bookpath, list, sheetname):
    print("INFO: {} updating with {}...".format(bookpath, sheetname))
    sheet = workbook.create_sheet(title=sheetname)

    for y in range(len(v)):
        for x in range(len(v[0])):
            list_value = list[y][x]
            sheet.cell(y + 1, x + 1, list_value)

    print("INFO: {} updated with {}".format(bookpath, sheetname))
    workbook.save(bookpath)


def formatcellstyles(workbook, bookpath):
    print("INFO: {} formatting cell styles...".format(bookpath))
    for sheet in workbook:
        for row_num in sheet.rows:
            row = row_num[0].row
            sheet.row_dimensions[row].height = 18

        for column_num in sheet.columns:
            column = column_num[0].column
            sheet.column_dimensions[get_column_letter(column)].width = 3.5

    workbook.save(bookpath)
    print("INFO: {} formated cell styles".format(bookpath))
    return


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    srcs = choosefiles()
    temp_min, temp_max = askminmax()

    wb = opxl.Workbook()
    time_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    dir = "./saved"
    if not os.path.exists(dir):
        os.makedirs(dir)
    bookpath = "./saved/runned_{}.xlsx".format(time_now)

    row_num, clumnn_num = 0, 0

    loaded_num = 0

    for src in srcs:
        if not (len(srcs) - loaded_num) == 1:
            print(
                "INFO: {} / {} files remaining".format(
                    (len(srcs) - loaded_num), len(srcs)
                )
            )
        else:
            print("INFO: 1 / {} file remaining".format(len(srcs)))

        filename = pathlib.Path(src).stem

        v = rdimg(src, temp_min, temp_max)
        row_num, clumnn_num = len(v), len(v[0])

        cvtsheet(wb, bookpath, v, filename)

        loaded_num += 1

    formatcellstyles(wb, bookpath)
    print("INFO: Saved to {}".format(bookpath))
    print("INFO: All operations have been completed")
    showmsgbox("convert-thermoshot", "{} に保存済み".format(bookpath))
